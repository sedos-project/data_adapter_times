import pandas as pd
import re
from openpyxl import Workbook, load_workbook
import pickle


def create_blank_excel(file_path):
    """
    Creates a blank Excel file with the specified path.

    Args:
        file_path (str): The path to the Excel file to be created.
    """
    wb = Workbook()
    wb.save(file_path)


def process_data(original_df: pd.DataFrame) -> pd.DataFrame:
    """
    This function takes in a pandas dataframe and processes the data to create a new dataframe with the specified columns.

    Args:
        original_df (pd.DataFrame): The original dataframe with the specified columns

    Returns:
        pd.DataFrame: The processed dataframe with the specified columns
    """
    technology_names = []
    comms_in = []
    comms_out = []
    attributes = []
    comm_grps = []
    comm_grp_elements = {}  # Dictionary to store comm_grps with their elements

    bracket_pattern = re.compile(r"\[([^]]+)\]")
    remove_pattern = re.compile(r"\b(pri_|sec_|iip_|exo_|emi_)")

    for _, row in original_df.iterrows():
        process = row.get("process", None)
        input_str = str(row.get("input", "")) if pd.notna(row.get("input")) else ""
        output_str = str(row.get("output", "")) if pd.notna(row.get("output")) else ""

        # Find and clean the bracketed items for the CommGrp value
        bracketed_items = bracket_pattern.findall(input_str)
        cleaned_bracketed_items = [
            remove_pattern.sub("", item)
            for bracketed_item in bracketed_items
            for item in bracketed_item.split(",")
        ]

        comm_grp_str = (
            "cg_" + "_".join(cleaned_bracketed_items) if cleaned_bracketed_items else ""
        )

        if bracketed_items:
            # Append ACT_EFF attribute and its CommGrp
            technology_names.append(process)
            comms_in.append(None)
            comms_out.append(None)
            attributes.append("ACT_EFF")
            comm_grps.append(comm_grp_str)

            # Append FLO_SHAR attributes for each bracketed item
            for original_item in bracketed_items:
                elements = [item.strip() for item in original_item.split(",")]
                for item in elements:
                    technology_names.append(process)
                    comms_in.append(item)
                    comms_out.append(None)
                    attributes.append("FLO_SHAR")
                    comm_grps.append(comm_grp_str)
                    # Collect elements for the comm_grp
                    if comm_grp_str in comm_grp_elements:
                        comm_grp_elements[comm_grp_str].add(item)
                    else:
                        comm_grp_elements[comm_grp_str] = set(elements)

        # Process the non-bracketed items normally
        for inp in re.sub(bracket_pattern, "", input_str).split(","):
            inp = inp.strip()
            if inp:  # Check if the input item is not an empty string after stripping
                technology_names.append(process)
                comms_in.append(inp)
                comms_out.append(None)
                attributes.append("INPUT")
                comm_grps.append("")

        for out in output_str.split(","):
            out = out.strip()
            if out:  # Check if the output item is not an empty string after stripping
                technology_names.append(process)
                comms_in.append(None)
                comms_out.append(out)
                attributes.append("OUTPUT")
                comm_grps.append("")

    # Make sure all lists are of the same length before creating the DataFrame
    if not (
        len(technology_names)
        == len(comms_in)
        == len(comms_out)
        == len(attributes)
        == len(comm_grps)
    ):
        raise ValueError("Lists are not of the same length, cannot form a DataFrame.")

    # Building the final DataFrame with specified columns
    data = {
        "TechName": technology_names,
        "TechDesc": ["" for _ in technology_names],
        "Attribute": [attr.upper() for attr in attributes],
        "Comm-IN": comms_in,
        "Comm-OUT": comms_out,
        "CommGrp": comm_grps,
        "TimeSlice": ["" for _ in technology_names],
        "LimType": ["" for _ in technology_names],
        "2021": ["" for _ in technology_names],
        "2024": ["" for _ in technology_names],
        "2027": ["" for _ in technology_names],
        "2030": ["" for _ in technology_names],
        "2035": ["" for _ in technology_names],
        "2040": ["" for _ in technology_names],
        "2045": ["" for _ in technology_names],
        "2050": ["" for _ in technology_names],
        "2060": ["" for _ in technology_names],
        "2070": ["" for _ in technology_names],
    }

    df = pd.DataFrame(data)

    # Sort the DataFrame by "TechName" and "Attribute"
    # Assigning a custom order for "Attribute"
    attribute_order = {"INPUT": 1, "OUTPUT": 2, "ACT_EFF": 3, "FLO_SHAR": 4}
    df["AttributeRank"] = df["Attribute"].map(attribute_order)
    df.sort_values(by=["TechName", "AttributeRank"], inplace=True)
    df.drop("AttributeRank", axis=1, inplace=True)

    # Convert set of elements to list for easier handling later
    for key in comm_grp_elements.keys():
        comm_grp_elements[key] = list(comm_grp_elements[key])

    return df, comm_grp_elements


def find_header_row(sheet, header_name):
    """
    This function finds the row number of the first occurrence of a specific header in a given sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet where the header is to be found.
    header_name (str): The name of the header to be found.

    Returns:
    int: The row number of the first occurrence of the header. If the header is not found, a ValueError is raised.

    Raises:
    ValueError: If the header is not found within the first 20 rows.
    """
    for row in range(1, 20):  # Assume headers are within the first 10 rows
        for col in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=row, column=col).value)
            if header_name.lower() in cell_value.lower():
                return row
    raise ValueError("Header row not found within the first 10 rows.")


def update_commodity_groups(file_path, comm_grps):
    """
    Updates the 'commodity_group' sheet in the Excel file located at 'file_path' with new commodity groups.

    Parameters:
    file_path (str): The path to the Excel file.
    comm_grps (dict): A dictionary where keys are commodity names and values are lists of elements associated with each commodity.

    Returns:
    None. The function modifies the existing workbook in-place.
    """
    # Load the workbook and select the 'commodity_group' sheet
    wb = load_workbook(file_path)
    ws = wb["commodity_group"]

    # Dynamically find the header row by looking for a known header, e.g., "Name"
    header_row = find_header_row(ws, "Name")
    name_col = None
    cset_cn_col = None

    # Scan the found header row to locate the correct columns based on header names
    for col in range(1, ws.max_column + 1):
        header_value = str(ws.cell(row=header_row, column=col).value)
        if header_value.strip().lower() == "name":
            name_col = col
        elif header_value.strip().lower() == "cset_cn":
            cset_cn_col = col

    # Validate that the necessary columns were found
    if not name_col or not cset_cn_col:
        raise ValueError("Required columns not found in the sheet")

    # Create a set to track existing commodity names for quick lookup
    existing_comms = set()
    # Start reading from the row just after the header row
    row_idx = header_row + 1
    while ws.cell(row=row_idx, column=name_col).value:
        existing_comms.add(ws.cell(row=row_idx, column=name_col).value.strip().lower())
        row_idx += 1

    # Reset row index to the next empty row
    while ws.cell(row=row_idx, column=name_col).value:
        row_idx += 1

    # Add new commodity groups to the sheet, checking for duplicates
    for comm_name, elements in comm_grps.items():
        # Check if the commodity name is already present
        if comm_name.strip().lower() not in existing_comms:
            ws.cell(row=row_idx, column=name_col).value = comm_name  # Name
            ws.cell(row=row_idx, column=cset_cn_col).value = ", ".join(
                elements
            )  # Cset_CN
            row_idx += 1

    # Save the workbook
    wb.save(file_path)


# Create the blank test_output.xlsx file
TIMES_FILE_PATH = "test_output.xlsx"
create_blank_excel(TIMES_FILE_PATH)

# Load the original DataFrame
SEDOS_FILE = pd.read_excel("test_data.xlsx")

# Process the data
times_df, commodity_groups = process_data(SEDOS_FILE)
print(times_df)

# Save the DataFrame to a pickle file
pickle_file_path = "times_df.pkl"
with open(pickle_file_path, "wb") as f:
    pickle.dump(times_df, f)

print(f"DataFrame saved to: {pickle_file_path}")

# Path to the SysSettings.xlsx
SYS_SETTINGS_PATH = "SysSettings.xlsx"

# Update the commodity groups in the SysSettings file
update_commodity_groups(SYS_SETTINGS_PATH, commodity_groups)
print(f"Updated Commodity Groups in: {SYS_SETTINGS_PATH}")
