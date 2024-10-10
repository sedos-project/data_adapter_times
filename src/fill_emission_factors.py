import pandas as pd
import requests
from openpyxl import load_workbook
import os

# Define the file path for output
SCEN_EMISSION_FILE_PATH = "output_data/Scen_emission_all.xlsx"
CO2_EMISSION_FILE_PATH = "output_data/vt_DE_CO2_Emission.xlsx"

# Initialize a set to track handled processes
handled_processes = set()

# Global variable for start_row
start_row = None

# Global variable for global_emission_data (fetched once at the beginning)
global_emission_data = None

# Data collection for batch processing
emi_co2_f_data = []


def fetch_data(url, process_name):
    """
    Fetches data from the API.
    """
    try:
        response = requests.get(url)
        if response.status_code == 200:
            print(f"Data fetched successfully for process: {process_name}")
            return pd.DataFrame(response.json())
        else:
            print(
                f"Failed to fetch data for process: {process_name}, status code: {response.status_code}"
            )
            return pd.DataFrame()
    except requests.RequestException as e:
        print(f"Error fetching data for process: {process_name}, error: {e}")
        return pd.DataFrame()


def process_emi_co2_f(other_indexes, remaining_string, global_emission_data):
    """
    Handles the logic when 'emi_co2_f_' is detected.
    Collects the data to be inserted later in batch.
    """
    if remaining_string in global_emission_data.columns:
        # Get the value from the 7th row in the global emission factors
        global_value = global_emission_data[remaining_string].iloc[6]
        # Collect data as a new column to be inserted later
        emi_co2_f_data.append(
            [
                other_indexes,
                remaining_string,
                "",  # Leave the units row unchanged
                global_value,
                global_value,
                global_value,
                global_value,
            ]
        )


def batch_insert_emi_co2_f_data():
    """
    Inserts all collected emi_co2_f_ data into the CO2 Emission Excel file in one go, ensuring unique combinations
    of `other_indexes` and `remaining_strings` are inserted.
    """
    # Load the CO2 Emission Excel file
    wb_co2 = load_workbook(CO2_EMISSION_FILE_PATH)
    if "emission" not in wb_co2.sheetnames:
        print("emission sheet not found.")
        return
    ws_co2 = wb_co2["emission"]

    # Find the CommName header and corresponding column
    header_row_co2 = find_header_row(ws_co2, "CommName")
    comm_name_col = None
    for col in range(1, ws_co2.max_column + 1):
        if ws_co2.cell(row=header_row_co2, column=col).value == "CommName":
            comm_name_col = col
            break

    if comm_name_col is None:
        print("CommName header not found.")
        return

    # Track unique combinations of other_indexes and remaining_strings
    unique_combinations = set()

    # Insert collected data as new columns
    start_col = comm_name_col + 1  # Start after the CommName column

    for data in emi_co2_f_data:
        # `data` contains [other_indexes, remaining_string, "", value1, value2, value3, value4]

        other_indexes = data[0]
        remaining_string = data[1]
        values = data[3:]  # The values to be inserted

        # Create a tuple for unique combination check
        combo = (other_indexes, remaining_string)

        # Only insert if this combination has not been inserted before
        if combo not in unique_combinations:
            # Add the combination to the set to prevent duplicates
            unique_combinations.add(combo)

            # Insert remaining string one row above the `other_indexes`
            ws_co2.cell(
                row=header_row_co2 - 1, column=start_col, value=remaining_string
            )

            # Insert the other_indexes on the row just below the remaining_string
            ws_co2.cell(row=header_row_co2, column=start_col, value=other_indexes)

            # Insert values in the next four rows below the `other_indexes`
            for i, value in enumerate(values):
                ws_co2.cell(row=header_row_co2 + 2 + i, column=start_col, value=value)

            # Move to the next column for the next batch of data
            start_col += 1

    # Save the CO2 Emission file after all modifications
    wb_co2.save(CO2_EMISSION_FILE_PATH)
    print(f"Batch insertion completed and saved to {CO2_EMISSION_FILE_PATH}.")


def find_header_row(sheet, header_name):
    """
    Finds the row number of the first occurrence of a specific header in a given sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet where the header is to be found.
    header_name (str): The name of the header to be found.

    Returns:
    int: The row number of the first occurrence of the header.
    """
    for row in range(1, 20):  # Assume headers are within the first 20 rows
        for col in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=row, column=col).value)
            if header_name.lower() in cell_value.lower():
                return row
    raise ValueError(f"Header row with '{header_name}' not found.")


def get_column_indices(sheet, header_row):
    """
    Get the column indices for specific columns in the Excel sheet based on the header row.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet containing the header.
    header_row (int): The row number where the headers are located.

    Returns:
    dict: A dictionary with column names as keys and their column indices as values.
    """
    headers = {cell.value: cell.column for cell in sheet[header_row]}
    required_columns = ["Attribute", "Other_Indexes", "Cset_CN", "Pset_PN", "DE"]
    if not all(col in headers for col in required_columns):
        raise ValueError("One or more required columns not found in the INS sheet.")

    return {
        "Attribute": headers["Attribute"],
        "Other_Indexes": headers["Other_Indexes"],
        "Cset_CN": headers["Cset_CN"],
        "Pset_PN": headers["Pset_PN"],
        "DE": headers["DE"],
    }


def clear_existing_data(ws, header_row):
    """
    Clears all the rows after the header row in the worksheet.

    Parameters:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to clear data from.
    header_row (int): The row number of the header.
    """
    print(f"Clearing data after header row {header_row}")
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.value = None  # Clear the cell value


def process_emission_factors(api_data, process_name, col_indices, ws):
    """
    Process emission factors from the API data and add them to the worksheet at the appropriate column positions.

    Parameters:
    api_data (pandas.DataFrame): The API data.
    process_name (str): The name of the process being handled.
    col_indices (dict): Column indices for "Attribute", "Other_Indexes", "Cset_CN", "Pset_PN", and "DE".
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to update.
    """
    global start_row  # Use a global start_row to keep appending correctly

    if api_data.empty:
        print(f"No emission factors found for process {process_name}")
        return

    for api_col in api_data.columns:
        if api_col.startswith("ef_"):
            # Split the column name to get Other_Indexes and Cset_CN
            col_parts = api_col.split("_emi_")
            if len(col_parts) != 2:
                continue  # Skip if column name format is not as expected

            other_indexes = col_parts[0].replace("ef_", "")
            cset_cn = "emi_" + col_parts[1]

            # Detect when 'emi_co2_f_' occurs
            if "emi_co2_f_" in cset_cn:
                # Get the value from the 7th row of the API data
                api_value = api_data[api_col].iloc[6]

                # Remove 'global_emission_factors.' and keep the remaining string
                if "global_emission_factors." in api_value:
                    remaining_string = api_value.replace("global_emission_factors.", "")
                    # Call the process_emi_co2_f function to handle this case, collecting data in memory
                    process_emi_co2_f(
                        other_indexes, remaining_string, global_emission_data
                    )

            else:
                # Determine the value to be pasted in the Attribute column
                attribute_value = (
                    "ENV_ACT" if "_p_" in cset_cn or "_proc_" in cset_cn else "FLO_EMIS"
                )

                # Get the value from the first row of the api_col
                api_value = (
                    api_data[api_col].iloc[6] if not api_data[api_col].empty else None
                )

                # If there is a value, add it to the worksheet
                if api_value is not None:
                    ws.cell(
                        row=start_row,
                        column=col_indices["Attribute"],
                        value=attribute_value,
                    )
                    ws.cell(
                        row=start_row,
                        column=col_indices["Other_Indexes"],
                        value=other_indexes,
                    )
                    ws.cell(row=start_row, column=col_indices["Cset_CN"], value=cset_cn)
                    ws.cell(
                        row=start_row, column=col_indices["Pset_PN"], value=process_name
                    )
                    ws.cell(row=start_row, column=col_indices["DE"], value=api_value)
                    start_row += 1
                else:
                    print(f"No value found for {api_col} in process {process_name}")


def process_group_or_individual(process_name, ws, is_group=False):
    """
    Processes emission data for a group or individual processes.
    Fetches API data and updates the Excel worksheet.

    Parameters:
    process_name (str): The name of the process or group.
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to be updated.
    is_group (bool): Whether the process_name represents a group or not.
    """
    # Skip if the process has already been handled
    if process_name in handled_processes:
        return

    # Fetch the API data for the process or group
    api_data = fetch_data(
        f"https://openenergy-platform.org/api/v0/schema/model_draft/tables/{process_name}/rows",
        process_name,
    )

    if api_data.empty:
        print(f"No data fetched for {process_name}. Skipping...")
        return  # Return early if no data is fetched

    # Mark the process as handled
    handled_processes.add(process_name)

    # Find the header row and column indices
    header_row = find_header_row(ws, "TimeSlice")
    col_indices = get_column_indices(ws, header_row)

    if is_group:
        # If it's a process group, handle the data by splitting on 'type'
        process_groups = api_data.groupby("type")
        for process, group_data in process_groups:
            if not process.endswith("_ag"):  # Skip processes ending with '_ag'
                # Clean the group_data by dropping columns where all values are empty
                cleaned_group_data = group_data.dropna(axis=1, how="all")
                process_emission_factors(cleaned_group_data, process, col_indices, ws)
                handled_processes.add(process)  # Mark each group process as handled
    else:
        # Handle individual processes
        process_emission_factors(api_data, process_name, col_indices, ws)


def main():
    global start_row  # Declare global so it can be used across functions
    global global_emission_data  # Declare global emission data variable

    # Fetch the global emission data once
    global_emission_url = "https://openenergyplatform.org/api/v0/schema/model_draft/tables/global_emission_factors/rows"

    try:
        response = requests.get(global_emission_url)
        if response.status_code == 200:
            print(f"Data fetched successfully for global emission factors")
            global_emission_data = pd.DataFrame(response.json())
        else:
            print(
                f"Failed to fetch data for global emission factors, status code: {response.status_code}"
            )
            global_emission_data = pd.DataFrame()
    except requests.RequestException as e:
        print(f"Error fetching data for global emission factors, error: {e}")
        global_emission_data = pd.DataFrame()

    if global_emission_data.empty:
        print("Global emission data could not be fetched. Exiting.")
        return

    # Load the existing workbook with openpyxl to read the existing structure
    wb = load_workbook(SCEN_EMISSION_FILE_PATH)

    if "INS" not in wb.sheetnames:
        print("INS sheet not found in the Excel file.")
        return

    ws = wb["INS"]  # Select the INS sheet

    # Find the header row and clear any existing data after the header
    header_row = find_header_row(ws, "TimeSlice")
    clear_existing_data(ws, header_row)
    start_row = header_row + 1  # Initialize start_row after clearing

    # Pre-defined process groups to handle
    process_groups = ["exo_other_ind"]  # Add more process groups as needed

    # Handle predefined process groups
    for process_group in process_groups:
        process_group_or_individual(process_group, ws, is_group=True)

    # Fetch and process data for individual processes (e.g., starting with 'ind')
    # Assuming times_df is pre-loaded or fetched from your initial pickle or Excel file
    times_df = pd.read_pickle("output_data/times_df_ind.pkl")
    unique_processes = times_df["TechName"].unique()
    ind_processes = [
        process for process in unique_processes if process.startswith("ind")
    ]

    # Skip processes that end with '_ag'
    ind_processes = [
        process for process in ind_processes if not process.endswith("_ag")
    ]

    for process in ind_processes:
        process_group_or_individual(process, ws)

    # Save the workbook after making updates
    wb.save(SCEN_EMISSION_FILE_PATH)
    print(f"Emission data processing completed and saved to {SCEN_EMISSION_FILE_PATH}.")

    # Perform the batch insert for emi_co2_f_ data
    batch_insert_emi_co2_f_data()


if __name__ == "__main__":
    main()
