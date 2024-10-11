import pandas as pd
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


def process_data(original_df: pd.DataFrame) -> pd.DataFrame:
    """
    This function takes in a pandas dataframe and processes the data to create a new dataframe with the specified columns.

    Args:
        original_df (pd.DataFrame): The original dataframe with the specified columns

    Returns:
        pd.DataFrame: The processed dataframe with the specified columns
    """
    technology_names = []
    comms_in = []
    comms_out = []
    attributes = []
    comm_grps = []
    comm_grp_elements = {}  # Dictionary to store comm_grps with their elements
    counter = 0
    bracket_pattern = re.compile(r"\[([^]]+)\]")
    remove_pattern = re.compile(r"\b(pri_|sec_|iip_|exo_|emi_)")

    for _, row in original_df.iterrows():
        process = row.get("process", None)
        if process is None or not process.lower().startswith("ind"):
            continue  # Skip this row if the process does not start with 'ind'

        if process.endswith("_ag"):
            continue  # Skip this row if the process ends with 'ag'

        input_str = str(row.get("input", "")) if pd.notna(row.get("input")) else ""
        output_str = str(row.get("output", "")) if pd.notna(row.get("output")) else ""

        counter += 1
        # Find and clean the bracketed items for the CommGrp value from input_str
        bracketed_items = bracket_pattern.findall(input_str)
        cleaned_bracketed_items = [
            remove_pattern.sub("", item)
            for bracketed_item in bracketed_items
            for item in bracketed_item.split(",")
        ]

        comm_grp_str = (
            "cg_" + "_".join([item[0] for item in cleaned_bracketed_items if item])
            if cleaned_bracketed_items
            else ""
        )

        if bracketed_items:
            # Append ACT_EFF attribute and its CommGrp
            technology_names.append(process)
            comms_in.append(None)
            comms_out.append(None)
            attributes.append("ACT_EFF")
            comm_grps.append(comm_grp_str)

            # Append FLO_SHAR attributes for each bracketed item
            for original_item in bracketed_items:
                elements = [item.strip() for item in original_item.split(",")]
                for item in elements:
                    technology_names.append(process)
                    comms_in.append(item)
                    comms_out.append(None)
                    attributes.append("FLO_SHAR")
                    comm_grps.append(comm_grp_str)
                    # Collect elements for the comm_grp
                    if comm_grp_str in comm_grp_elements:
                        comm_grp_elements[comm_grp_str].add(item)
                    else:
                        comm_grp_elements[comm_grp_str] = set(elements)

        # Now process the output_str for bracketed items, similar to input_str
        output_bracketed_items = bracket_pattern.findall(output_str)
        cleaned_output_bracketed_items = [
            remove_pattern.sub("", item)
            for bracketed_item in output_bracketed_items
            for item in bracketed_item.split(",")
        ]

        comm_grp_str_out = (
            "cg_"
            + "_".join([item[0] for item in cleaned_output_bracketed_items if item])
            if cleaned_output_bracketed_items
            else ""
        )

        if output_bracketed_items:
            # Append FLO_SHAR attributes for each bracketed item in output_str
            for original_item in output_bracketed_items:
                elements = [item.strip() for item in original_item.split(",")]
                for item in elements:
                    technology_names.append(process)
                    comms_in.append(None)
                    comms_out.append(item)
                    attributes.append("OUTPUT")
                    comm_grps.append(None)
                    # Collect elements for the comm_grp
                    if comm_grp_str_out in comm_grp_elements:
                        comm_grp_elements[comm_grp_str_out].add(item)
                    else:
                        comm_grp_elements[comm_grp_str_out] = set(elements)

        # Process the non-bracketed items in input_str normally
        for inp in re.sub(bracket_pattern, "", input_str).split(","):
            inp = inp.strip()
            if inp:  # Check if the input item is not an empty string after stripping
                technology_names.append(process)
                comms_in.append(inp)
                comms_out.append(None)
                attributes.append("INPUT")
                comm_grps.append("")

        # Process the non-bracketed items in output_str
        for out in re.sub(bracket_pattern, "", output_str).split(","):
            out = out.strip()
            if out:  # Check if the output item is not an empty string after stripping
                technology_names.append(process)
                comms_in.append(None)
                comms_out.append(out)
                attributes.append("OUTPUT")
                comm_grps.append("")

    # Make sure all lists are of the same length before creating the DataFrame
    if not (
        len(technology_names)
        == len(comms_in)
        == len(comms_out)
        == len(attributes)
        == len(comm_grps)
    ):
        raise ValueError("Lists are not of the same length, cannot form a DataFrame.")

    # Building the final DataFrame with specified columns
    data = {
        "TechName": technology_names,
        "TechDesc": ["" for _ in technology_names],
        "Attribute": [attr.upper() for attr in attributes],
        "Comm-IN": comms_in,
        "Comm-OUT": comms_out,
        "CommGrp": comm_grps,
        "TimeSlice": ["" for _ in technology_names],
        "LimType": ["" for _ in technology_names],
        "2021": ["" for _ in technology_names],
        "2024": ["" for _ in technology_names],
        "2027": ["" for _ in technology_names],
        "2030": ["" for _ in technology_names],
        "2035": ["" for _ in technology_names],
        "2040": ["" for _ in technology_names],
        "2045": ["" for _ in technology_names],
        "2050": ["" for _ in technology_names],
        "2060": ["" for _ in technology_names],
        "2070": ["" for _ in technology_names],
    }

    df = pd.DataFrame(data)

    # Sort the DataFrame by "TechName" and "Attribute"
    # Assigning a custom order for "Attribute"
    attribute_order = {"INPUT": 1, "OUTPUT": 2, "ACT_EFF": 3, "FLO_SHAR": 4}
    df["AttributeRank"] = df["Attribute"].map(attribute_order)
    df.sort_values(by=["TechName", "AttributeRank"], inplace=True)
    df.drop("AttributeRank", axis=1, inplace=True)

    # Convert set of elements to list for easier handling later
    for key in comm_grp_elements.keys():
        comm_grp_elements[key] = list(comm_grp_elements[key])

    print(f"Process counter: {counter}")
    return df, comm_grp_elements


def add_comm_sheet_to_workbook(file_path, processed_df):
    """
    This function adds a new sheet named 'Commodities' to an existing Excel file.
    It populates the sheet with commodity data and sets their membership in commodity sets.

    Parameters:
    file_path (str): The path to the existing Excel file.
    processed_df (DataFrame): The DataFrame containing the processed data.

    Returns:
    None. The function modifies the existing Excel file in-place.
    """
    # Load the existing workbook
    wb = load_workbook(file_path)

    # Create a new sheet
    ws_comm = wb.create_sheet("Commodity List")

    # Define fills, fonts, borders, and alignment
    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    blue_font = Font(color="0000FF", size=11, bold=True)
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Set ~FI_Comm title
    ws_comm["B1"] = "~FI_Comm"
    ws_comm["B1"].font = blue_font
    ws_comm["B1"].alignment = align_center

    subheader_font = Font(color="000000")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Define headers and subheaders
    headers = [
        "Csets",
        "CommName",
        "CommDesc",
        "Unit",
        "LimType",
        "CTSLvl",
        "PeakTS",
        "Ctype",
    ]

    subheaders = [
        "I: Commodity Set Membership",
        "*Commodity Name",
        "Commodity Description",
        "Unit",
        "Balance Equ Type Override",
        "Timeslice Tracking Level",
        "Peak Monitoring",
        "Electricity Indicator",
    ]

    # Write headers and subheaders
    for col, header in enumerate(headers, start=2):  # Start from the second column
        cell = ws_comm.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = align_center

    for col, subheader in enumerate(
        subheaders, start=2
    ):  # Start from the second column
        cell = ws_comm.cell(row=3, column=col)
        cell.value = subheader
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = align_center

    # Load the commodity_set data from mapping_v3.xlsx
    wb_mapping = load_workbook("config_data/mapping_v3.xlsx", data_only=True)
    ws_mapping = wb_mapping["commodity_set"]

    # Find the header row dynamically
    header_row = find_header_row(ws_mapping, "CommName")

    # Create a dictionary to store commodity set memberships
    commodity_set_dict = {}
    for row in ws_mapping.iter_rows(min_row=header_row + 1, values_only=True):
        comm_name = str(row[0]).lower().strip() if row[0] else ""
        cset = (
            str(row[3]).strip() if row[3] else ""
        )  # Adjust the index based on the actual column for Csets
        commodity_set_dict[comm_name] = cset

    # Determine the commodity set membership
    commodity_sets = {}
    for commodity in set(
        processed_df["Comm-IN"].dropna().unique().tolist()
        + processed_df["Comm-OUT"].dropna().unique().tolist()
    ):
        commodity_lower = commodity.lower()
        if commodity_lower.startswith("exo_"):
            commodity_sets[commodity] = "DEM"
        elif commodity_lower.startswith("emi_"):
            commodity_sets[commodity] = "ENV"
        elif (
            commodity_lower in commodity_set_dict
            and commodity_set_dict[commodity_lower] == "MAT"
        ):
            commodity_sets[commodity] = "MAT"
        else:
            commodity_sets[commodity] = "NRG"

    # Populate the CommName column with unique commodities and set membership
    for row_idx, comm in enumerate(
        commodity_sets.keys(), start=4
    ):  # Start from the fourth row
        ws_comm.cell(row=row_idx, column=2, value=commodity_sets[comm])  # Csets
        ws_comm.cell(row=row_idx, column=3, value=comm)  # CommName

    # Adjust column widths
    for col in ws_comm.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 1  # Add 1 padding and adjust for a better fit
        ws_comm.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(file_path)


def add_process_sheet_to_workbook(file_path, processed_df):
    """
    This function adds a new sheet named 'Processes' to an existing Excel workbook.
    The sheet is populated with process-related data based on the input DataFrame.

    Parameters:
    file_path (str): The path to the existing Excel workbook.
    processed_df (DataFrame): A DataFrame containing processed data.

    Returns:
    None. The function modifies the existing workbook in-place.
    """
    # Load the existing workbook
    wb = load_workbook(file_path)

    # Create a new sheet
    ws_process = wb.create_sheet("Process List")

    # Define fills, fonts, borders, and alignment for headers and subheaders
    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    blue_font = Font(color="0000FF", size=11, bold=True)
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Set ~FI_Process title
    ws_process["B1"] = "~FI_Process"
    ws_process["B1"].font = blue_font
    ws_process["B1"].alignment = align_center

    subheader_font = Font(color="000000")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Define headers and subheaders
    headers = [
        "Sets",
        "TechName",
        "TechDesc",
        "Tact",
        "Tcap",
        "Tslvl",
        "PrimaryCG",
        "Vintage",
    ]

    subheaders = [
        "I: Process Set Membership",
        "*Technology Name",
        "Technology Description",
        "Activity Unit",
        "Capacity Unit",
        "Timeslice Operational Level",
        "Operational Commodity Group",
        "Vintage Tracking",
    ]

    # Write headers and subheaders
    for col, header in enumerate(headers, start=2):  # Start from the second column
        cell = ws_process.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = align_center

    for col, subheader in enumerate(
        subheaders, start=2
    ):  # Start from the second column
        cell = ws_process.cell(row=3, column=col)
        cell.value = subheader
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = align_center

    # Determine the process set membership
    process_sets = {}
    for _, row in processed_df.iterrows():
        tech_name = row["TechName"]
        output_commodities = row["Comm-OUT"]
        if "_chp_" in tech_name.lower():
            process_sets[tech_name] = "CHP"
        elif pd.notna(output_commodities) and "exo_" in output_commodities.lower():
            process_sets[tech_name] = "DEM"
        else:
            process_sets.setdefault(tech_name, "PRE")

    # Populate the process sheet
    for row_idx, tech_name in enumerate(process_sets.keys(), start=4):
        ws_process.cell(row=row_idx, column=2, value=process_sets[tech_name])  # Sets
        ws_process.cell(row=row_idx, column=3, value=tech_name)  # TechName

    # Adjust column widths
    for col in ws_process.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 1  # Add padding and adjust for a better fit
        ws_process.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(file_path)


def find_header_row(sheet, header_name):
    """
    This function finds the row number of the first occurrence of a specific header in a given sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet where the header is to be found.
    header_name (str): The name of the header to be found.

    Returns:
    int: The row number of the first occurrence of the header. If the header is not found, a ValueError is raised.

    Raises:
    ValueError: If the header is not found within the first 20 rows.
    """
    for row in range(1, 20):  # Assume headers are within the first 10 rows
        for col in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=row, column=col).value)
            if header_name.lower() in cell_value.lower():
                return row
    raise ValueError("Header row not found within the first 10 rows.")


def update_commodity_groups(file_path, comm_grps):
    """
    Updates the 'commodity_group' sheet in the Excel file located at 'file_path' with new commodity groups.

    Parameters:
    file_path (str): The path to the Excel file.
    comm_grps (dict): A dictionary where keys are commodity names and values are lists of elements associated with each commodity.

    Returns:
    None. The function modifies the existing workbook in-place.
    """
    # Load the workbook and select the 'commodity_group' sheet
    wb = load_workbook(file_path)
    ws = wb["commodity_group"]

    # Dynamically find the header row by looking for a known header, e.g., "Name"
    header_row = find_header_row(ws, "Name")
    name_col = None
    cset_cn_col = None
    all_regions_col = None  # New column for 'AllRegions'

    # Scan the found header row to locate the correct columns based on header names
    for col in range(1, ws.max_column + 1):
        header_value = str(ws.cell(row=header_row, column=col).value)
        if header_value.strip().lower() == "name":
            name_col = col
        elif header_value.strip().lower() == "cset_cn":
            cset_cn_col = col
        elif header_value.strip().lower() == "allregions":  # Find AllRegions column
            all_regions_col = col

    # Validate that the necessary columns were found
    if not name_col or not cset_cn_col or not all_regions_col:
        raise ValueError("Required columns not found in the sheet")

    # Create a set to track existing commodity names for quick lookup
    existing_comms = set()
    # Start reading from the row just after the header row
    row_idx = header_row + 1
    while ws.cell(row=row_idx, column=name_col).value:
        existing_comms.add(ws.cell(row=row_idx, column=name_col).value.strip().lower())
        row_idx += 1

    # Reset row index to the next empty row
    while ws.cell(row=row_idx, column=name_col).value:
        row_idx += 1

    # Add new commodity groups to the sheet, checking for duplicates
    for comm_name, elements in comm_grps.items():
        # Check if the commodity name is already present
        if comm_name.strip().lower() not in existing_comms:
            ws.cell(row=row_idx, column=name_col).value = comm_name  # Name
            ws.cell(row=row_idx, column=cset_cn_col).value = ", ".join(
                elements
            )  # Cset_CN
            ws.cell(row=row_idx, column=all_regions_col).value = (
                "y"  # AllRegions set to 'y'
            )
            row_idx += 1

    # Save the workbook
    wb.save(file_path)


def create_blank_excel(file_path):
    """
    Creates a blank Excel file with the necessary structure for the subsequent operations.

    Parameters:
    file_path (str): The path where the blank Excel file will be created.

    Returns:
    None. The function creates an Excel file at the specified path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Process Data"
    wb.save(file_path)


def filter_output_with_emi_commodities(df: pd.DataFrame) -> pd.DataFrame:
    """
    This function filters out rows from the DataFrame where the Attribute is 'OUTPUT'
    and the Comm-OUT starts with 'emi_'.

    Args:
        df (pd.DataFrame): The original DataFrame.

    Returns:
        pd.DataFrame: The filtered DataFrame.
    """
    # Filter out rows where 'Attribute' is 'OUTPUT' and 'Comm-OUT' starts with 'emi_'
    filtered_df = df[
        ~((df["Attribute"] == "OUTPUT") & df["Comm-OUT"].str.startswith("emi_"))
    ].copy()

    return filtered_df


# Load the original DataFrame
SEDOS_FILE = pd.read_excel("input_data/Modellstruktur.xlsx", sheet_name="Process_Set")

# Process the data
times_df, commodity_groups = process_data(SEDOS_FILE)
print(times_df)

# Apply the filter function to remove the rows where attribute is 'OUTPUT' and comm-out starts with 'emi_'
times_df_filtered = filter_output_with_emi_commodities(times_df)

# Define the path for the pickle file
PICKLE_FILE_PATH = "output_data/times_df_ind.pkl"

# Save the filtered times_df DataFrame as a pickle file
times_df_filtered.to_pickle(PICKLE_FILE_PATH)
print(f"Filtered times_df DataFrame saved as pickle file: {PICKLE_FILE_PATH}")

# Path to the SysSettings.xlsx
SYS_SETTINGS_PATH = "config_data/SysSettings.xlsx"
# Update the commodity groups in the SysSettings file
update_commodity_groups(SYS_SETTINGS_PATH, commodity_groups)
print(f"Updated Commodity Groups in: {SYS_SETTINGS_PATH}")

# Format and save the Excel file
TIMES_FILE_PATH = "output_data/vt_DE_ind.xlsx"
create_blank_excel(TIMES_FILE_PATH)
add_comm_sheet_to_workbook(TIMES_FILE_PATH, times_df)
add_process_sheet_to_workbook(TIMES_FILE_PATH, times_df)

print("Excel file saved")
