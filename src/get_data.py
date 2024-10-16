import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Initialize the counter
fetch_data_counter = 0
units_mapping = {}


def format_and_save_excel(file_path, processed_df):
    """
    Format and save the processed data into an Excel file.

    Parameters:
    processed_df (pandas.DataFrame): The DataFrame containing the processed data.
    file_path (str): The path where the Excel file will be saved.

    Returns:
    str: The path where the Excel file is saved.
    """

    # Replace pd.NA with empty strings
    processed_df = processed_df.fillna("")

    wb = load_workbook(file_path)
    ws = wb.active

    # Existing setup for fills, fonts, borders
    # Define two color fills for alternating rows
    fill1 = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")
    fill2 = PatternFill(start_color="C5d9F1", end_color="C5d9F1", fill_type="solid")
    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )  # Assuming light green color for subheaders
    header_font = Font(name="Arial", size=10, bold=True, color="000000")
    sub_header_font = Font(name="Arial", size=10, bold=False, color="000000")
    table_name_font = Font(name="Arial", size=12, bold=True, color="0000FF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    no_border = Border()

    def style_cell(cell, fill=None, font=None, border=None, alignment=None):
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        else:
            cell.font = sub_header_font
        if border is not None:  # Only apply the border if it is explicitly given
            cell.border = border
        if alignment:
            cell.alignment = alignment

    # Headers and subheaders for the Excel sheet
    headers = [
        "",
        "TechName",
        "*TechDesc",
        "Attribute",
        "Comm-IN",
        "Comm-OUT",
        "CommGrp",
        "TimeSlice",
        "LimType",
        "2021",
        "2024",
        "2027",
        "2030",
        "2035",
        "2040",
        "2045",
        "2050",
        "2060",
        "2070",
    ]

    subheaders = [
        "",
        "*Technology Name",
        "Technology Description",
        "Attribute Declaration\nColumn",
        "Input\nCommodity",
        "Output\nCommodity",
        "Commodity\nGroup",
        "Time Slices\ndefinition",
        "Bound\ndefinition",
        "Base\nYear",
        "Data\nYear",
        "Data\nYear",
        "Data\nYear",
        "Data\nYear",
        "Data\nYear",
        "Data\nYear",
        "Data\nYear",
        "Data\nYear",
        "Data\nYear",
    ]

    # Write the table name with blue font
    table_name_cell = ws.cell(row=1, column=2, value="~FI_T")
    style_cell(
        table_name_cell, font=table_name_font, alignment=Alignment(horizontal="center")
    )

    # Write the headers with style
    for col, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header_title)
        style_cell(
            cell,
            fill=header_fill,
            font=header_font,
            border=thin_border,
            alignment=Alignment(horizontal="center", wrap_text=True),
        )

    # Write the subheaders with style
    for col, sub_header_title in enumerate(subheaders, start=1):
        cell = ws.cell(row=3, column=col, value=sub_header_title)
        style_cell(
            cell,
            fill=subheader_fill,
            font=sub_header_font,
            border=thin_border,
            alignment=Alignment(horizontal="center", wrap_text=True),
        )

    # Calculate column widths based on headers and subheaders
    column_widths = [
        max(len(header), max(len(part) for part in subheader.split("\n")))
        for header, subheader in zip(headers, subheaders)
    ]

    # Initialize the variable to keep track of the current process and the fill to apply
    current_process = None
    current_fill = fill1

    # Write the data and format the cells with alternating colors
    for row_index, (idx, row) in enumerate(
        processed_df.iterrows(), start=4
    ):  # Data starts from row 4
        process = row["TechName"]
        if process != current_process:
            # Switch the fill when the process changes
            current_fill = fill2 if current_fill == fill1 else fill1
            current_process = process

        for col_index, (col, value) in enumerate(row.items(), start=2):
            # Convert empty lists to empty strings
            if isinstance(value, list) and not value:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(map(str, value))
            cell = ws.cell(row=row_index, column=col_index, value=value)
            style_cell(cell, fill=current_fill, border=no_border)
            # Update the max length if the current value is longer
            column_widths[col_index - 1] = max(
                column_widths[col_index - 1], len(str(value))
            )

    # Set column widths with a little extra padding
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width + 1

    # Save the workbook
    wb.save(file_path)
    return file_path


def data_units(metadata):
    """
    Converts the units of the fields in the API data according to the metadata.

    Parameters:
    api_data (pandas.DataFrame): The DataFrame containing the fetched API data.
    metadata (dict): The metadata containing information about the units.

    Returns:
    pandas.DataFrame: The DataFrame with units converted to the standard units.
    """
    units_dict = {}  # Initialize an empty dictionary to store units

    if not metadata:
        return units_dict  # Return as is if no data or metadata is present

    for resource in metadata.get("resources", []):
        fields = resource.get("schema", {}).get("fields", [])
        for field in fields:
            field_name = field["name"]
            field_unit = field.get("unit")
            if field_name and field_unit:
                units_dict[field_name] = field_unit  # Store the field_name and unit

    return units_dict


def update_commodity_list_units(excel_file_path, units_mapping):
    """
    Update the 'Commodity List' sheet in the Excel file with the units from units_mapping.
    """
    from openpyxl import load_workbook

    # Load the Excel file
    wb = load_workbook(excel_file_path)
    if "Commodity List" in wb.sheetnames:
        ws = wb["Commodity List"]
    else:
        print("Commodity List sheet not found in the Excel file.")
        return

    # Get header row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            if cell.value == "CommName":
                header_row = cell.row
                break
        if header_row is not None:
            break

    if header_row is None:
        print("CommName header not found in Commodity List sheet.")
        return

    # Get column indices for CommName and Unit
    headers = {cell.value: cell.column for cell in ws[header_row]}
    if "CommName" in headers and "Unit" in headers:
        commname_col = headers["CommName"]
        unit_col = headers["Unit"]
    else:
        print("CommName or Unit column not found in Commodity List sheet.")
        return

    # For each field_name and field_unit in units_mapping
    for field_name, field_unit in units_mapping.items():
        # Only process field names starting with 'conversion_factor_'
        if field_name.startswith("conversion_factor_"):
            # Remove the 'conversion_factor_' prefix directly
            comm_name = field_name[len("conversion_factor_") :]
            found = False
            for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
                commname_cell = row[commname_col - 1]  # openpyxl columns are 1-based
                if commname_cell.value and isinstance(commname_cell.value, str):
                    if commname_cell.value.strip().lower() == comm_name.strip().lower():
                        unit_cell = row[unit_col - 1]
                        unit_cell.value = field_unit
                        found = True
                        break  # Assuming CommName is unique
            if not found:
                print(
                    f"CommName '{comm_name}' with unit '{field_unit}' not found in Commodity List sheet."
                )

    # Save the workbook
    wb.save(excel_file_path)
    print("Commodity List sheet updated with units.")


def fetch_data(url, process_name):
    global fetch_data_counter
    fetch_data_counter += 1  # Increment the counter
    try:
        response = requests.get(url)
        if response.status_code == 200:
            print(
                f"Data fetched successfully for process {fetch_data_counter}: {process_name}"
            )
            return pd.DataFrame(response.json())
        else:
            print(
                f"Failed to fetch data for process {fetch_data_counter}: {process_name}, status code: {response.status_code}"
            )
            return pd.DataFrame()  # Return an empty DataFrame if status code is not 200
    except requests.RequestException as e:
        print(
            f"No data found for process {fetch_data_counter}: {process_name}, error: {e}"
        )
        return pd.DataFrame()  # Return an empty DataFrame in case of error


def fetch_process_metadata(process):
    """
    Fetches the metadata of a process and extracts the units for all the resources.

    Parameters:
    process_name (str): The name of the process to fetch metadata for.

    Returns:
    dict: A dictionary where the keys are the resource names and the values are their corresponding units.
    """
    try:
        url = f"https://openenergy-platform.org/api/v0/schema/model_draft/tables/{process}/meta/"
        response = requests.get(url)
        if response.status_code == 200:
            metadata = response.json()
            # print(f"Units fetched successfully for process: {process}")
            return metadata
        else:
            # print(
            #     f"Failed to fetch metadata for process: {process}, status code: {response.status_code}"
            # )
            return {}
    except requests.RequestException as e:
        print(f"Error fetching metadata for process: {process}, error: {e}")
        return {}


def find_header_row(sheet, header_name):
    """
    This function finds the row number of the first occurrence of a specific header in a given sheet.

    Parameters:
    sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet where the header is to be found.
    header_name (str): The name of the header to be found.

    Returns:
    int: The row number of the first occurrence of the header. If the header is not found, a ValueError is raised.

    Raises:
    ValueError: If the header is not found within the first 20 rows.
    """
    for row in range(1, 20):  # Assume headers are within the first 10 rows
        for col in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=row, column=col).value)
            if header_name.lower() in cell_value.lower():
                return row
    raise ValueError("Header row not found within the first 10 rows.")


def data_mapping(times_df, process_name, is_group=False):
    """
    Fetches data from the API for a given process name or group and updates the times_df DataFrame.

    Parameters:
    times_df (pandas.DataFrame): The DataFrame containing the initial data.
    process_name (str): The name of the process or process group to fetch and process data for.
    is_group (bool): Flag indicating whether the process_name is a group.

    Returns:
    pandas.DataFrame: The updated DataFrame with the new data merged.
    """
    api_process_data = fetch_data(
        f"https://openenergy-platform.org/api/v0/schema/model_draft/tables/{process_name}/rows",
        process_name,
    )

    if api_process_data.empty:
        return times_df  # Return the original DataFrame if no data is fetched

    if is_group:
        # Divide the data based on the 'type' column
        process_groups = api_process_data.groupby("type")

        process_count = 0  # Initialize a counter for the processes handled

        for process, group_data in process_groups:
            if process.endswith("_ag"):  # Skip processes ending with _ag
                continue
            handled_processes.append(process)
            times_df = data_mapping_internal(
                times_df, process, group_data
            )  # Call internal function for each process
            process_count += 1  # Increment the counter for each handled process

        print(
            f"{process_count} processes were handled inside the process group: {process_name}"
        )
        return times_df
    else:
        return data_mapping_internal(times_df, process_name, api_process_data)


def data_mapping_internal(times_df, process_name, api_process_data):

    # Fetch metadata
    metadata = fetch_process_metadata(process_name)

    # Update units_mapping
    global units_mapping
    units_mapping.update(data_units(metadata))

    # Filter for the specific process and keep track of the index range
    times_df_filtered = times_df[times_df["TechName"] == process_name]
    if times_df_filtered.empty:
        print(f"{process_name} was not found in the input and hence was skipped")
        return times_df  # Skip if there is no matching process

    start_idx = times_df.index.get_loc(times_df_filtered.index[0])
    end_idx = times_df.index.get_loc(times_df_filtered.index[-1])

    # Load the mapping file
    mapping_file_path = "config_data/mapping_v3.xlsx"
    wb = load_workbook(mapping_file_path, data_only=True)
    sheet = wb["SEDOS_parameters"]

    # Find the header row for 'SEDOS'
    header_row = find_header_row(sheet, "SEDOS")

    # Extract the SEDOS, TIMES, and Constraints columns
    sedos_list = []
    times_list = []
    constraints_list = []

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        sedos_value = row[0].value  # Assuming SEDOS is in the first column
        times_value = row[1].value  # Assuming TIMES is in the second column
        constraints_value = row[8].value  # Assuming Constraints is in the ninth column
        if sedos_value and times_value:
            sedos_list.append(sedos_value)
            times_list.append(times_value)
            constraints_list.append(
                constraints_value if constraints_value is not None else ""
            )

    # Modify the SEDOS list items
    sedos_list = [item.split("<")[0].lower().strip() for item in sedos_list]

    # Create a mapping dictionary
    mapping_dict = dict(zip(sedos_list, times_list))
    constraints_dict = dict(zip(sedos_list, constraints_list))

    # Create a dictionary for matched SEDOS items and API column names
    matched_columns = {}

    for sedos_item in sedos_list:
        for api_col in api_process_data.columns:
            if sedos_item in api_col.lower():
                if sedos_item not in matched_columns:
                    matched_columns[sedos_item] = []
                matched_columns[sedos_item].append(api_col)

    # Add the TIMES list items and constraints corresponding to the matched SEDOS items
    extended_matched_columns = {
        sedos_item: (api_cols, mapping_dict[sedos_item], constraints_dict[sedos_item])
        for sedos_item, api_cols in matched_columns.items()
    }

    # Update the times_df_filtered with the api_process_data based on the matched columns
    for sedos_item, (
        api_cols,
        times_col,
        constraint,
    ) in extended_matched_columns.items():
        for api_col in api_cols:
            if api_col in api_process_data.columns:
                # Extract the values and year from the API data
                api_values = api_process_data[api_col]
                years = api_process_data["year"]
                comm_col_value = api_col.replace("conversion_factor_", "")

                for api_value, year in zip(api_values, years):
                    # Find the column in times_df_filtered that matches the year
                    if str(year) in times_df_filtered.columns:
                        # Check if sedos_item contains 'conversion_factor_'
                        if "conversion_factor_" in sedos_item:
                            # Check if both the Attribute and Comm-IN/Comm-OUT match
                            matching_row = times_df_filtered[
                                (
                                    (times_df_filtered["Attribute"] == times_col)
                                    | (times_df_filtered["Attribute"] == "OUTPUT")
                                    | (times_df_filtered["Attribute"] == "INPUT")
                                )
                                & (
                                    (times_df_filtered["Comm-IN"] == comm_col_value)
                                    | (times_df_filtered["Comm-OUT"] == comm_col_value)
                                )
                            ]
                            if not matching_row.empty:
                                for idx in matching_row.index:
                                    if api_value is not None:
                                        times_df_filtered.at[idx, str(year)] = api_value
                            else:
                                matching_row = times_df_filtered[
                                    (times_df_filtered["Attribute"] == "ACT_EFF")
                                ]
                                if not matching_row.empty:
                                    for idx in matching_row.index:
                                        if api_value is not None:
                                            times_df_filtered.at[idx, str(year)] = (
                                                1 / api_value
                                            )
                        elif "flow_share" in sedos_item:
                            # Add flow share values
                            matching_row = times_df_filtered[
                                times_df_filtered["Attribute"] == times_col
                            ]
                            if not matching_row.empty:
                                comm_in_out_values = []
                                for idx in matching_row.index:
                                    comm_in_out_values.extend(
                                        [
                                            times_df_filtered.at[idx, "Comm-IN"],
                                            times_df_filtered.at[idx, "Comm-OUT"],
                                        ]
                                    )
                                comm_in_out_values = list(
                                    filter(pd.notna, comm_in_out_values)
                                )
                                # Parse the current api_col to get flow share commodity
                                flow_share_commodity = api_col.replace(
                                    sedos_item, ""
                                ).strip("_")

                                # Add values to the matching rows
                                sum_of_matched_values = 0
                                for idx in matching_row.index:
                                    comm_in = times_df_filtered.at[idx, "Comm-IN"]
                                    comm_out = times_df_filtered.at[idx, "Comm-OUT"]
                                    if flow_share_commodity in (comm_in, comm_out):
                                        if api_value is not None:
                                            times_df_filtered.at[idx, str(year)] = (
                                                api_value / 100
                                            )
                                            times_df_filtered.at[idx, "LimType"] = (
                                                constraint
                                            )
                                            sum_of_matched_values += api_value / 100

                        elif (
                            "availability_constant" in sedos_item
                            or "availability_timeseries_fixed" in sedos_item
                        ):
                            # Handle availability constants or time series fixed
                            matching_row = times_df_filtered[
                                times_df_filtered["Attribute"] == times_col
                            ]
                            if matching_row.empty:
                                # Add a new row if the Attribute does not exist
                                new_row = pd.Series(
                                    {col: pd.NA for col in times_df_filtered.columns}
                                )
                                new_row["TechName"] = process_name
                                new_row["Attribute"] = times_col
                                new_row["LimType"] = constraint
                                times_df_filtered = pd.concat(
                                    [times_df_filtered, new_row.to_frame().T],
                                    ignore_index=True,
                                )
                                new_row_idx = times_df_filtered[
                                    times_df_filtered["Attribute"] == times_col
                                ].index[-1]
                                if api_value is not None:
                                    times_df_filtered.at[new_row_idx, str(year)] = (
                                        api_value / 100
                                    )
                            else:
                                for idx in matching_row.index:
                                    if api_value is not None:
                                        times_df_filtered.at[idx, str(year)] = (
                                            api_value / 100
                                        )
                                        times_df_filtered.at[idx, "LimType"] = (
                                            constraint
                                        )
                        else:
                            # Check if only the Attribute matches
                            matching_row = times_df_filtered[
                                times_df_filtered["Attribute"] == times_col
                            ]
                            if matching_row.empty:
                                # Add a new row if the Attribute does not exist
                                new_row = pd.Series(
                                    {col: pd.NA for col in times_df_filtered.columns}
                                )
                                new_row["TechName"] = process_name
                                new_row["Attribute"] = times_col
                                new_row["LimType"] = constraint
                                times_df_filtered = pd.concat(
                                    [times_df_filtered, new_row.to_frame().T],
                                    ignore_index=True,
                                )
                                new_row_idx = times_df_filtered[
                                    times_df_filtered["Attribute"] == times_col
                                ].index[-1]
                                if api_value is not None:
                                    # If the sedos_item contains 'cb_coefficient', apply 1/api_value
                                    if "cb_coefficient" in sedos_item:
                                        times_df_filtered.at[new_row_idx, str(year)] = (
                                            1 / api_value
                                        )
                                    else:
                                        # For all other cases, just use the api_value directly
                                        times_df_filtered.at[new_row_idx, str(year)] = (
                                            api_value
                                        )
                            else:
                                for idx in matching_row.index:
                                    if api_value is not None:
                                        times_df_filtered.at[idx, str(year)] = api_value
                                        times_df_filtered.at[idx, "LimType"] = (
                                            constraint
                                        )

    # Implement CAP2ACT logic
    cap2act_value = pd.NA  # Default to empty if metadata is not fetched

    # Fetch metadata
    metadata = fetch_process_metadata(process_name)

    if metadata:  # Check if metadata was successfully fetched
        cap2act_value = 1  # Default CAP2ACT value if metadata is present

        if process_name.endswith("_1"):
            # Check if process has 'cost_in_p' field in schema->fields and if its unit is GW
            resources = metadata.get("resources", [])
            for resource in resources:
                fields = resource.get("schema", {}).get("fields", [])
                for field in fields:
                    if field["name"] == "cost_inv_p" and field.get("unit") == "M€/GW":
                        cap2act_value = 31.536
                        break
        elif process_name.endswith("_0"):
            # Check if process has 'capacity_p_inst_0' field in schema->fields and if its unit is GW
            resources = metadata.get("resources", [])
            for resource in resources:
                fields = resource.get("schema", {}).get("fields", [])
                for field in fields:
                    if (
                        field["name"] == "capacity_p_inst_0"
                        and field.get("unit") == "GW"
                    ):
                        cap2act_value = 31.536
                        break

    # Add CAP2ACT as a new row in times_df_filtered
    cap2act_row = pd.Series(
        {
            "TechName": process_name,
            "Attribute": "CAP2ACT",
            "LimType": pd.NA,
            "2021": cap2act_value,
            "2024": cap2act_value,
            "2027": cap2act_value,
            "2030": cap2act_value,
            "2035": cap2act_value,
            "2040": cap2act_value,
            "2045": cap2act_value,
            "2050": cap2act_value,
            "2060": cap2act_value,
            "2070": cap2act_value,
        }
    )
    times_df_filtered = pd.concat(
        [times_df_filtered, cap2act_row.to_frame().T], ignore_index=True
    )

    # Replace <NA> with empty strings before updating the original times_df
    with pd.option_context("future.no_silent_downcasting", True):
        times_df_filtered = times_df_filtered.fillna("")

    # Ensure the updated times_df_filtered has the same or larger index range
    if len(times_df_filtered) > (end_idx - start_idx + 1):
        # Split the original times_df into three parts
        before = times_df.iloc[:start_idx]
        after = times_df.iloc[end_idx + 1 :]

        # Concatenate the before part, updated times_df_filtered, and the after part
        times_df = pd.concat([before, times_df_filtered, after], ignore_index=True)
    else:
        times_df.iloc[start_idx : end_idx + 1] = times_df_filtered.values

    return times_df


# Paths and URLs
TIMES_FILE_PATH = "output_data/vt_DE_ind.xlsx"

# Read the pickle file and print the DataFrame
PICKLE_FILE_PATH = "output_data/times_df_ind.pkl"
times_df = pd.read_pickle(PICKLE_FILE_PATH)

# Create a copy of times_df to work with
updated_df = times_df.copy()

# Pre-defined process groups to handle
process_groups = [
    "exo_other_ind",  # Add other process groups here if needed
]

# Define a global list to keep track of processes that have been handled
handled_processes = []

# Handle pre-defined process groups first
for process_group in process_groups:
    updated_df = data_mapping(updated_df, process_group, is_group=True)

# Fetch and process data for each unique process in the TechName column that starts with 'ind'
unique_processes = times_df["TechName"].unique()
ind_processes = [process for process in unique_processes if process.startswith("ind")]

# Skip processes that end with '_ag'
ind_processes = [process for process in ind_processes if not process.endswith("_ag")]

for process in ind_processes:
    if process not in handled_processes:
        updated_df = data_mapping(
            updated_df, process
        )  # Perform data mapping and update updated_df

format_and_save_excel(TIMES_FILE_PATH, updated_df)
update_commodity_list_units(TIMES_FILE_PATH, units_mapping)
print("Excel file saved")
